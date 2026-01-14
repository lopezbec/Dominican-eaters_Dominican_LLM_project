"""Universal Excel formatter for all content types with professional styling."""

from typing import List, Dict, Any
import logging
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class AvailabilityColors:
    """
    Color scheme for availability status.
    
    Single Responsibility: Define color constants.
    """
    FOUND = {'fill': 'C6EFCE', 'font': '006100'}
    PARTIAL = {'fill': 'FFEB9C', 'font': '9C5700'}
    NOT_FOUND = {'fill': 'FFC7CE', 'font': '9C0006'}


class ExcelFormatter:
    """
    Universal Excel formatter with configurable styling.
    
    Single Responsibility: Excel file creation and formatting.
    """
    
    def __init__(
        self,
        sheet_name: str = "Datos",
        header_color: str = "366092",
        header_font_color: str = "FFFFFF"
    ):
        self.sheet_name = sheet_name
        self.header_color = header_color
        self.header_font_color = header_font_color
    
    def save_to_excel(
        self,
        items: List[Any],
        filename: str,
        column_widths: Dict[str, int],
        center_align_columns: List[str] = None,
        availability_column: str = None,
        wrap_text_columns: List[str] = None
    ) -> bool:
        """
        Save items to Excel with formatting.
        
        Args:
            items: List of objects with to_dict() method
            filename: Output filename
            column_widths: Dict mapping column letters to widths
            center_align_columns: List of column letters to center-align
            availability_column: Column letter for availability color-coding
            wrap_text_columns: List of column letters to wrap text
            
        Returns:
            True if successful, False otherwise
        """
        try:
            data = [item.to_dict() for item in items]
            df = pd.DataFrame(data)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=self.sheet_name)
                worksheet = writer.sheets[self.sheet_name]
                
                self._apply_column_widths(worksheet, column_widths)
                self._apply_header_style(worksheet)
                self._apply_data_formatting(
                    worksheet,
                    items,
                    center_align_columns or [],
                    wrap_text_columns or []
                )
                
                if availability_column:
                    self._apply_availability_colors(
                        worksheet,
                        items,
                        availability_column,
                        df.columns.get_loc(self._get_column_name_from_letter(df, availability_column)) + 1
                    )
            
            logger.info("Excel saved successfully: %s", filename)
            return True
            
        except Exception as e:
            logger.error("Error saving Excel %s: %s", filename, e)
            return False
    
    def _apply_column_widths(self, worksheet: Worksheet, widths: Dict[str, int]):
        """
        Apply column width settings.
        
        Single Responsibility: Column width configuration.
        """
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width
    
    def _apply_header_style(self, worksheet: Worksheet):
        """
        Apply styling to header row.
        
        Single Responsibility: Header formatting.
        """
        header_fill = PatternFill(
            start_color=self.header_color,
            end_color=self.header_color,
            fill_type='solid'
        )
        header_font = Font(bold=True, color=self.header_font_color)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _apply_data_formatting(
        self,
        worksheet: Worksheet,
        items: List[Any],
        center_columns: List[str],
        wrap_columns: List[str]
    ):
        """
        Apply formatting to data rows.
        
        Single Responsibility: Data cell formatting.
        """
        for row in range(2, len(items) + 2):
            for col in center_columns:
                cell = worksheet[f"{col}{row}"]
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for col in wrap_columns:
                cell = worksheet[f"{col}{row}"]
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def _apply_availability_colors(
        self,
        worksheet: Worksheet,
        items: List[Any],
        column_letter: str,
        column_index: int
    ):
        """
        Apply color-coding based on availability status.
        
        Single Responsibility: Availability color application.
        """
        for row_idx, item in enumerate(items, start=2):
            cell = worksheet.cell(row=row_idx, column=column_index)
            
            availability = getattr(item, 'disponibilidad', None)
            
            if availability == "ENCONTRADO":
                colors = AvailabilityColors.FOUND
            elif availability == "PARCIAL":
                colors = AvailabilityColors.PARTIAL
            else:
                colors = AvailabilityColors.NOT_FOUND
            
            cell.fill = PatternFill(
                start_color=colors['fill'],
                end_color=colors['fill'],
                fill_type='solid'
            )
            cell.font = Font(color=colors['font'])
    
    def _get_column_name_from_letter(self, df: pd.DataFrame, letter: str) -> str:
        """
        Get column name from Excel letter (A=0, B=1, etc.).
        
        Single Responsibility: Column letter to index conversion.
        """
        col_idx = ord(letter.upper()) - ord('A')
        return df.columns[col_idx]


def save_books_to_excel(books: List[Any], filename: str) -> bool:
    """
    Factory function for saving books with proper formatting.
    
    Single Responsibility: Books-specific Excel configuration.
    """
    formatter = ExcelFormatter(sheet_name="Audiolibros Dominicanos")
    
    column_widths = {
        'A': 10,
        'B': 40,
        'C': 30,
        'D': 10,
        'E': 60,
        'F': 15,
        'G': 25,
        'H': 15,
        'I': 80
    }
    
    return formatter.save_to_excel(
        items=books,
        filename=filename,
        column_widths=column_widths,
        center_align_columns=['A', 'D', 'F', 'H'],
        availability_column='H',
        wrap_text_columns=['I']
    )


def save_songs_to_excel(songs: List[Any], filename: str) -> bool:
    """
    Factory function for saving songs with proper formatting.
    
    Single Responsibility: Songs-specific Excel configuration.
    """
    formatter = ExcelFormatter(sheet_name="Canciones")
    
    column_widths = {
        'A': 20,
        'B': 25,
        'C': 30,
        'D': 80,
        'E': 50,
        'F': 50,
        'G': 25,
    }
    
    return formatter.save_to_excel(
        items=songs,
        filename=filename,
        column_widths=column_widths,
        wrap_text_columns=['D']
    )


def save_poems_to_excel(poems: List[Any], filename: str) -> bool:
    """
    Factory function for saving poems with proper formatting.
    
    Single Responsibility: Poems-specific Excel configuration.
    """
    formatter = ExcelFormatter(sheet_name="Poemas Dominicanos")
    
    column_widths = {
        'A': 10,
        'B': 40,
        'C': 30,
        'D': 10,
        'E': 20,
        'F': 60,
        'G': 15,
        'H': 25,
        'I': 20,
        'J': 15,
        'K': 40,
        'L': 15,
        'M': 80
    }
    
    return formatter.save_to_excel(
        items=poems,
        filename=filename,
        column_widths=column_widths,
        center_align_columns=['A', 'D', 'G', 'I', 'J', 'L'],
        availability_column='L',
        wrap_text_columns=['M']
    )
