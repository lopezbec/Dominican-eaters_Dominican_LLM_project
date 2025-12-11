"""Main orchestration service for scraping books from librosdominicanos.com."""

import logging
import os
from pathlib import Path
from typing import List
import pandas as pd
from tqdm import tqdm

from ..clients.librosdominicanos_client import LibrosDominicanosScraper
from ..clients.youtube_client import YouTubeClient
from ..services.pdf_processor import PDFProcessor
from ..models.scraped_book import ScrapedBook

logger = logging.getLogger(__name__)


class BookScraperService:
    """
    Main service for orchestrating the book scraping process.
    Follows SOLID principles with clear separation of concerns.
    """
    
    def __init__(
        self,
        output_dir: str = "scraped_books",
        download_pdfs: bool = True,
        search_youtube: bool = True
    ):
        """
        Initialize the book scraper service.
        
        Args:
            output_dir: Directory to save downloaded PDFs
            download_pdfs: Whether to download PDF files
            search_youtube: Whether to search for YouTube audiobooks
        """
        self.output_dir = Path(output_dir)
        self.pdf_dir = self.output_dir / "pdfs"
        self.download_pdfs = download_pdfs
        self.search_youtube = search_youtube
        
        self.scraper = LibrosDominicanosScraper(delay=1.0)
        self.youtube_client = YouTubeClient(videos_per_search=3)
        self.pdf_processor = PDFProcessor()
        
        os.makedirs(self.pdf_dir, exist_ok=True)
        
        logger.info(f"Initialized BookScraperService with output_dir: {self.output_dir}")
    
    def scrape_all_books(self) -> List[ScrapedBook]:
        """
        Scrape all books from librosdominicanos.com.
        
        Returns:
            List of ScrapedBook objects with all collected data
        """
        logger.info("Starting book scraping process")
        
        logger.info("Step 1: Scraping book listings")
        book_listings = self.scraper.scrape_all_listings()
        logger.info(f"Found {len(book_listings)} books")
        
        scraped_books = []
        
        logger.info("Step 2: Processing each book")
        for idx, book_info in enumerate(tqdm(book_listings, desc="Processing books"), 1):
            try:
                scraped_book = self._process_single_book(idx, book_info)
                scraped_books.append(scraped_book)
                
            except Exception as e:
                logger.error(f"Error processing book {idx}: {e}")
                scraped_book = ScrapedBook(
                    book_number=idx,
                    title=book_info.get('title', 'Unknown'),
                    author=book_info.get('author', 'Unknown'),
                    year=book_info.get('year', 'N/A'),
                    book_url=book_info.get('book_url', '')
                )
                scraped_books.append(scraped_book)
        
        logger.info(f"Completed scraping {len(scraped_books)} books")
        return scraped_books
    
    def _process_single_book(self, book_number: int, book_info: dict) -> ScrapedBook:
        """
        Process a single book: scrape details, download PDF, search YouTube.
        
        Args:
            book_number: Sequential book number
            book_info: Dictionary with basic book info
            
        Returns:
            ScrapedBook object with all data
        """
        title = book_info['title']
        author = book_info['author']
        year = book_info['year']
        book_url = book_info['book_url']
        
        scraped_book = ScrapedBook(
            book_number=book_number,
            title=title,
            author=author,
            year=year,
            book_url=book_url
        )
        
        detail_info = self.scraper.scrape_book_detail_page(book_url)
        if detail_info:
            scraped_book.pdf_url = detail_info.get('pdf_url')
        
        if self.download_pdfs and scraped_book.pdf_url:
            pdf_filename = f"book_{book_number:03d}.pdf"
            pdf_path = self.pdf_dir / pdf_filename
            
            downloaded_path = self.scraper.download_pdf(scraped_book.pdf_url, str(pdf_path))
            if downloaded_path:
                scraped_book.mark_pdf_downloaded()
                
                markdown_content = self.pdf_processor.pdf_to_markdown(downloaded_path)
                if markdown_content:
                    scraped_book.content = markdown_content
        
        if self.search_youtube:
            youtube_info = self.youtube_client.search_audiobook(title, author)
            if youtube_info:
                scraped_book.set_youtube_info(
                    url=youtube_info['url'],
                    duration=youtube_info['duration'],
                    content_type=youtube_info['type']
                )
        
        return scraped_book
    
    def save_to_excel(self, books: List[ScrapedBook], filename: str = "book_content_scrape.xlsx") -> bool:
        """
        Save scraped books to Excel file with proper formatting.
        
        Args:
            books: List of ScrapedBook objects
            filename: Output filename
            
        Returns:
            True if successful, False otherwise
        """
        try:
            output_path = self.output_dir / filename
            
            data = [book.to_dict() for book in books]
            df = pd.DataFrame(data)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Books')
                
                worksheet = writer.sheets['Books']
                
                column_widths = {
                    'A': 12,  # Book Number
                    'B': 45,  # Title
                    'C': 30,  # Author
                    'D': 10,  # Year
                    'E': 50,  # PDF URL
                    'F': 15,  # PDF Downloaded
                    'G': 80,  # Content
                    'H': 50,  # YouTube URL
                    'I': 15,  # YouTube Duration
                    'J': 25,  # YouTube Content Type
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                for row in range(2, len(books) + 2):
                    worksheet.row_dimensions[row].height = 60
                
                from openpyxl.styles import Alignment, Font, PatternFill
                
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for row in range(2, len(books) + 2):
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'H', 'I', 'J']:
                        cell = worksheet[f'{col}{row}']
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    content_cell = worksheet[f'G{row}']
                    content_cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    pdf_downloaded = worksheet[f'F{row}'].value
                    if pdf_downloaded == "Yes":
                        worksheet[f'F{row}'].fill = PatternFill(
                            start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'
                        )
                        worksheet[f'F{row}'].font = Font(color='006100')
                    else:
                        worksheet[f'F{row}'].fill = PatternFill(
                            start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
                        )
                        worksheet[f'F{row}'].font = Font(color='9C0006')
                    
                    youtube_url = worksheet[f'H{row}'].value
                    if youtube_url and youtube_url != "NO ENCONTRADO":
                        worksheet[f'H{row}'].fill = PatternFill(
                            start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'
                        )
                        worksheet[f'H{row}'].font = Font(color='006100')
                    else:
                        worksheet[f'H{row}'].fill = PatternFill(
                            start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
                        )
                        worksheet[f'H{row}'].font = Font(color='9C0006')
            
            logger.info(f"Excel file saved successfully: {output_path}")
            print(f"\n✓ Excel file saved: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            print(f"\n✗ Error saving Excel file: {e}")
            return False
    
    def generate_statistics(self, books: List[ScrapedBook]) -> dict:
        """
        Generate statistics about the scraped books.
        
        Args:
            books: List of ScrapedBook objects
            
        Returns:
            Dictionary with statistics
        """
        stats = {
            'total_books': len(books),
            'pdfs_found': sum(1 for b in books if b.pdf_url),
            'pdfs_downloaded': sum(1 for b in books if b.pdf_downloaded == "Yes"),
            'youtube_found': sum(1 for b in books if b.youtube_url != "NO ENCONTRADO"),
            'content_extracted': sum(1 for b in books if b.content)
        }
        
        return stats
